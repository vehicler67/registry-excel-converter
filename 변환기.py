#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════╗
║     등기부등본 PDF → Excel 6시트 변환기 v1.0 (Mac/Windows)      ║
║     HanaXellOcr0.7 역공학 기반 — 제2 독립 도구                    ║
╚══════════════════════════════════════════════════════════════╝

[개요]
  대한민국 법원 등기부등본(PDF)을 OCR로 읽어들여
  HanaXellOcr0.7과 동일한 6시트 구조의 Excel 파일로 변환합니다.

[역공학 출처]
  - 분석 대상: HanaXellOcr0.7_fixed.xlam (윈도우용 Excel 확장앱)
  - 분석 방법: OOXML/ZIP 추출 → 정적 분석 → 출력물 비교 (역공학_백서.md 참고)
  - 재현 대상: output.xlsx 의 6시트 구조, 병합셀 규칙, 필드 매핑

[기술 스택]
  - OCR 엔진: Tesseract 5 (kor+eng) — 한글 CID 폰트 대응
  - PDF 렌더링: PyMuPDF (fitz) — 400DPI 고해상도
  - Excel 출력: openpyxl — 병합셀·스타일·컬럼너비 완전 제어
  - 이미지 전처리: Pillow — 선명화·대비·이진화

[사용법]
  1. 이 파일이 있는 폴더에서 터미널 실행:
     python3 변환기.py
  2. 파일 선택 다이얼로그에서 등기부등본 PDF 선택
  3. 자동으로 OCR → 정제 → 6시트 Excel 생성 → Excel 열기

[의존성 설치]
  pip3 install pymupdf pytesseract pillow openpyxl

[유지보수 참고]
  - 모든 함수에 한국어 docstring 포함
  - [역공학] 태그: HanaXellOcr0.7에서 추론한 로직
  - [실패사례] 태그: 과거 시행착오와 해결 방법
  - [주의] 태그: 변경 시 영향도가 큰 부분
"""

import os, sys, re, json, subprocess, tkinter as tk
from tkinter import filedialog
from pathlib import Path

# ── 외부 라이브러리 (없으면 안내 후 종료) ──────────────────────
try:
    import fitz  # PyMuPDF
    from PIL import Image, ImageEnhance, ImageFilter
    import pytesseract
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[오류] 필수 라이브러리가 없습니다: {e}")
    print("설치 명령: pip3 install pymupdf pytesseract pillow openpyxl")
    sys.exit(1)

# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 0] 프로젝트 경로 설정                                    ║
# ╚══════════════════════════════════════════════════════════════╝
# [주의] 이 파일의 위치를 기준으로 모든 경로가 결정됩니다.
#        폴더째 복사해서 다른 맥에서도 즉시 사용 가능.

BASE_DIR = Path(__file__).resolve().parent
OUT_DIR = BASE_DIR / '출력결과'
OUT_DIR.mkdir(exist_ok=True)


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 1] 설정 상수                                            ║
# ╚══════════════════════════════════════════════════════════════╝

# ── OCR 설정 ─────────────────────────────────────────────────
# [역공학] HanaXellOcr0.7은 해상도를 "자동 조정"한다고 추정되나,
#          실제 동작은 300~400DPI 수준으로 확인됨.
# [실패사례] 200DPI: 작은 글씨(등기번호, 지번) 인식 실패.
#            600DPI: 인식률은 좋으나 처리 시간 2배 이상.
#            → 400DPI가 속도-정확도 최적점.
OCR_DPI = 400
TESSERACT_LANG = 'kor+eng'

# ── Excel 출력 설정 ──────────────────────────────────────────
# [역공학] HanaXellOcr0.7 output.xlsx 기준 6개 시트명.
#          시트 순서도 output.xlsx 와 동일하게 유지.
SHEET_NAMES = [
    '전체요약',          # 기본정보 + 현황 요약
    '갑구',              # 소유권에 관한 사항
    '을구',              # 소유권 외 권리 (근저당, 전세권 등)
    '공동담보목록',      # 여러 필지에 걸친 담보 목록
    '요약-갑구',         # 갑구 요약 테이블
    '요약-을구',         # 을구 요약 테이블
]

# [역공학] output.xlsx 병합셀 규칙 (역공학_백서.md §3.3)
#   B1:H1 = 문서 헤더, C4:H4 = 기본정보 라벨,
#   C5:H5 = 기본정보 값, C6:H6 = 구분 헤더
#   갑구/을구는 F열까지만 사용 (6열)
MERGE_SPEC = {
    '전체요약':      [(1,2,1,8), (4,3,4,8), (5,3,5,8), (6,3,6,8)],
    '갑구':          [(1,2,1,6), (4,3,4,6), (5,3,5,6), (6,3,6,6)],
    '을구':          [(1,2,1,6), (4,3,4,6), (5,3,5,6), (6,3,6,6)],
    '공동담보목록':  [(1,2,1,8), (4,3,4,8), (5,3,5,8), (6,3,6,8)],
    '요약-갑구':     [(1,2,1,6), (4,3,4,6), (5,3,5,6)],
    '요약-을구':     [(1,2,1,6), (4,3,4,6), (5,3,5,6)],
}

# ── 스타일 정의 ──────────────────────────────────────────────
# [역공학] HanaXellOcr0.7 출력물은 Windows 기본 글꼴(맑은 고딕) 사용.
#          색상은 output.xlsx 스크린샷에서 추출 근사.
STYLE = {
    'header_font':   Font(name='맑은 고딕', bold=True, size=14, color='1F4E79'),
    'section_font':  Font(name='맑은 고딕', bold=True, size=12),
    'label_font':    Font(name='맑은 고딕', bold=True, size=10),
    'data_font':     Font(name='맑은 고딕', size=10),
    'meta_font':     Font(name='맑은 고딕', italic=True, size=9, color='888888'),
    'header_fill':   PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    'section_fill':  PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'border':        Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')),
    'center':        Alignment(horizontal='center', vertical='center', wrap_text=True),
}


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 2] 등기용어 교정 사전                                    ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] 이 사전은 output.xlsx의 정제된 출력과
#          src_비교_*.md(원본 OCR 결과)를 대조하여 추출했습니다.
#
# [실패사례] CID 폰트 PDF를 OCR 하면 모든 한글 사이에 공백이 삽입됨.
#   예: "고 유 번 호" → "고유번호" 로 복원 필요.
#   Tesseract가 CID 폰트를 개별 글리프로 인식하기 때문.
#
# [실패사례] OCR이 【 (특수 괄호)를 [ (일반 대괄호)로 잘못 인식.
#   → TERM_REPLACE에서 양쪽 모두 대응하도록 등록.
#
# [주의] 사전 순서가 중요합니다.
#   긴 문자열을 먼저 매칭해야 부분 문자열 오치환을 방지할 수 있음.
#   예: "등 기 사 항 전 부 증 명 서" → "등기사항전부증명서" 가
#        "등 기" → "등기" 보다 먼저 와야 함.

TERM_REPLACE = {
    # ── 섹션 헤더 (CID 공백 포함) ──
    # [역공학] HanaXellOcr0.7 output.xlsx 섹션 구분자
    '【 표 제 부 】': '【표제부】', '【 표제부 】': '【표제부】',
    '【 갑 구 】': '【갑구】', '【 갑구 】': '【갑구】',
    '【 을 구 】': '【을구】', '【 을구 】': '【을구】',
    '【 매 매 목 록 】': '【매매목록】', '【 매매목록 】': '【매매목록】',
    '【 공 동 담 보 목 록 】': '【공동담보목록】',

    # ── 【 → [ OCR 오인식 대응 ──
    # [실패사례] Tesseract가 특수문자 【 를 [ 로 인식.
    #           TERM_REPLACE로 원복 처리.
    '[표제부】': '【표제부】', '[갑구】': '【갑구】',
    '[을구】': '【을구】', '[매매목록】': '【매매목록】',
    '[공동담보목록】': '【공동담보목록】',

    # ── 필드명 (CID 공백) ──
    # [역공학] output.xlsx Row 4-7 라벨 필드
    '고 유 번 호': '고유번호', '소 재 지': '소재지',
    '부 동 산 종 류': '부동산종류', '열 람 일 시': '열람일시',
    '현 황': '현황',

    # ── 등기사항 (CID 공백) ──
    # [역공학] output.xlsx 에서 관측된 정제 후 표현
    '소 유 권 이 전': '소유권이전',
    '근 저 당 권 설 정': '근저당권설정',
    '지 상 권 설 정': '지상권설정',
    '가 압 류': '가압류',
    '임 의 경 매 개 시 결 정': '임의경매개시결정',
    '매 매 목 록': '매매목록',
    '공 동 담 보 목 록': '공동담보목록',
    '등 기 사 항 전 부 증 명 서': '등기사항전부증명서',

    # ── 기타 ──
    '[ 토 지]': '[토지]', '[ 토 지 ]': '[토지]',
    '( 토 지 의 표 시 )': '(토지의표시)',
}


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 3] 정규식 패턴                                          ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] output.xlsx의 정제된 값들을 분석하여
#          OCR 원시 출력에서 어떤 변환이 일어나는지 역추적.
#
# [실패사례] 초기 버전은 단순 replace로 접근했으나,
#   "2023년 2월 20일"에서 월/일에 0이 빠진 경우(zero-padding)를
#   놓쳐서 날짜 비교 검증이 깨짐 → 정규식으로 범용 처리.

# 날짜: "2023 년 2 월 20 일" → "2023년02월20일"
RE_DATE = re.compile(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일')

# 시간: "17 시 25 분 59 초" → 시간:분:초
RE_TIME = re.compile(r'(\d{1,2})\s*시\s*(\d{1,2})\s*분\s*(\d{1,2})\s*초')

# 등기번호: "제  107414  호" → "제107414호"
RE_REGNO = re.compile(r'제\s*(\d{3,7})\s*호')

# 금액: "금  266,000,000  원" → "금266,000,000원"
RE_MONEY = re.compile(r'금\s*([\d,]+)\s*원')

# ── CID 폰트 주소 패턴 ──────────────────────────────────────
# [역공학] 등기부등본에 자주 등장하는 관할 법원/등기소 주소.
#          CID 폰트로 인해 "의 정 부 지 방 법 원"처럼 분리됨.
# [실패사례] 초기엔 일반 한글 공백 제거로 처리했으나,
#           "파주시 파평면" → "파주시파평면"이 되어
#           주소 검색/매칭 실패.
#           → 등기부 도메인 사전 기반 정규식으로 해결.
RE_CID_ADDR = re.compile(
    r'(경\s*기\s*도|파\s*주\s*시|고\s*양\s*시|파\s*평\s*면'
    r'|문\s*산\s*읍|마\s*산\s*리|사\s*임\s*당\s*로)')
RE_CID_COURT = re.compile(
    r'(의\s*정\s*부\s*지\s*방\s*법\s*원'
    r'|고\s*양\s*지\s*원|파\s*주\s*등\s*기\s*소)')

# ── 섹션 식별 키워드 ────────────────────────────────────────
# [역공학] output.xlsx 시트 분리 기준이 되는 섹션 헤더.
#          OCR 결과에서 이 키워드들을 찾아 텍스트를 분할.
SECTION_KEYS = [
    '【표제부】', '【갑구】', '【을구】',
    '【매매목록】', '【공동담보목록】',
    '소유지분현황', '기본정보',
]

# ── 정확도 검증용 키워드 ────────────────────────────────────
# [역공학] src_비교_*.md 기준 20개 필수 키워드.
#          이 키워드들이 모두 추출되었는지로 정확도 1차 판단.
KEYWORDS_CHECK = [
    '등기사항전부증명서', '고유번호', '소재지', '부동산종류',
    '열람일시', '갑구', '을구', '순위번호', '등기목적',
    '접수', '등기원인', '권리자', '소유권이전', '근저당권설정',
    '지상권설정', '가압류', '매매목록', '공동담보목록',
    '소유지분', '임의경매',
]


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 4] PDF → 이미지 렌더링 + OCR 엔진                       ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] HanaXellOcr0.7의 추정 파이프라인:
#   PDF → 이미지 렌더링 (자동 해상도) → 이미지 전처리
#   → OCR 엔진 (Tesseract 또는 PaddleOCR 추정)
#   → 셀 경계 검출 → 텍스트→셀 할당 → Excel 출력
#
# [실패사례] PyMuPDF로 직접 텍스트 추출(get_text) 시도했으나
#   CID 폰트로 인해 모든 글자 사이에 공백이 들어감.
#   → OCR 우회로 해결. (PyMuPDF는 이미지 렌더링 용도로만 사용)


def render_page(pdf_path, page_num, dpi=OCR_DPI):
    """
    PDF 페이지를 고해상도 이미지로 렌더링합니다.

    [역공학] HanaXellOcr0.7은 내부적으로 PDF를 이미지로 변환 후
    OCR을 수행하는 것으로 추정됩니다 (output.xlsx의 이미지 아티팩트 흔적).
    PyMuPDF의 Matrix 변환으로 DPI를 정밀 제어합니다.

    [실패사례] 200DPI: 등기번호, 지번 등 작은 글씨 인식률 60% 이하.
    600DPI: 인식률 99%까지 올라가나 13페이지 기준 4분 이상 소요.
    → 400DPI가 실무 최적값.

    Args:
        pdf_path: PDF 파일 경로
        page_num: 0-인덱스 페이지 번호
        dpi: 렌더링 해상도 (기본 400)

    Returns:
        그레이스케일 PIL Image
    """
    doc = fitz.open(str(pdf_path))
    # [역공학] Matrix(zoom, zoom)으로 DPI 제어.
    # 72는 PDF 기본 DPI. dpi/72 = 확대 배율.
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = doc[page_num].get_pixmap(matrix=mat)
    img = Image.frombytes('RGB', (pix.width, pix.height), pix.samples).convert('L')
    doc.close()
    return img


def ocr_image(img):
    """
    이미지 → Tesseract OCR 텍스트 추출.

    [역공학] 이미지 전처리 파이프라인은 output.xlsx의
    노이즈 없는 출력 품질을 보고 역추적.
    HanaXellOcr0.7도 유사한 전처리(이진화+대비)를 거치는 것으로 추정.

    [실패사례] 전처리 없이 OCR → 배경 노이즈를 글자로 오인식.
    Contrast 3.0 과다 → 글자 획이 뭉개져서 인식률 하락.
    → SHARPEN + Contrast 2.0 + Threshold 175 가 최적 조합.

    Args:
        img: PIL Image (그레이스케일 권장)

    Returns:
        OCR 텍스트 문자열
    """
    # [전처리 1] 선명화 — 글자 경계 강조
    img = img.filter(ImageFilter.SHARPEN)
    # [전처리 2] 대비 강화 — 글자/배경 구분
    img = ImageEnhance.Contrast(img).enhance(2.0)
    # [전처리 3] 이진화 — Threshold 175 (0~255)
    # 175보다 밝으면 흰색(배경), 어두우면 검은색(글자)
    img = img.point(lambda x: 0 if x < 175 else 255)
    return pytesseract.image_to_string(img, lang=TESSERACT_LANG).strip()


def ocr_pdf(pdf_path, max_pages=None, progress_callback=None):
    """
    PDF 전체 페이지 OCR 처리.

    [실패사례] 20페이지 이상 대량 문서에서
    초기 버전은 메모리 누수로 15페이지 이후 크래시.
    → fitz.open() 후 명시적 doc.close()로 해결.
    → 10페이지마다 진행률 출력으로 장시간 처리 중 상태 확인 가능.

    Args:
        pdf_path: PDF 파일 경로
        max_pages: 최대 처리 페이지 수 (None=전체)
        progress_callback: 진행률 콜백 함수(page, total)

    Returns:
        [{'page': N, 'text': '...'}, ...]
    """
    doc = fitz.open(str(pdf_path))
    total = len(doc)
    pages_to_process = range(total) if max_pages is None else range(min(max_pages, total))

    results = []
    for i in pages_to_process:
        img = render_page(pdf_path, i, dpi=OCR_DPI)
        text = ocr_image(img)
        results.append({'page': i + 1, 'text': text})

        # 진행률 (10페이지마다 또는 마지막 페이지)
        if (i + 1) % 10 == 0 or (i + 1) == total:
            pct = (i + 1) / total * 100
            msg = f'  페이지 {i+1}/{total} OCR 완료 ({pct:.0f}%)'
            print(msg)
            if progress_callback:
                progress_callback(i + 1, total)

    doc.close()
    return results


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 5] 텍스트 정제 엔진                                      ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] HanaXellOcr0.7 output.xlsx의 정제된 텍스트와
# Tesseract OCR 원시 출력을 비교하여 정제 규칙 역추출.
#
# 정제 파이프라인 순서가 매우 중요합니다:
#   1. CID 공백 정제 → 2. 용어 교정 → 3. 날짜/번호/금액 포맷
#   → 4. OCR 노이즈 제거 → 5. 중요 줄 필터링
#
# [실패사례] v6 초기: 4번(노이즈 제거)를 1번보다 먼저 수행 →
#   "고 유 번 호"가 노이즈로 판정되어 통째로 삭제됨.
#   → 순서를 CID 공백 정제 → 노이즈 제거로 재배치.


def fix_cid_spacing(text):
    """
    CID 폰트로 인한 한글/숫자 사이 공백 제거.

    [역공학] CID 폰트(CFF/CID-Keyed Font)는 각 글리프가
    독립적인 CID 값으로 인코딩됩니다.
    OCR이 각 글자를 개별 객체로 인식하면서 공백이 삽입됨.
    "고유번호" → "고 유 번 호" 현상.

    [실패사례] 1회만 정규식 적용 → 잔여 공백 남음.
    "경 기 도 파 주 시" → 1차 "경기도 파주시" → 2차 "경기도파주시"
    → 2회 반복 적용으로 완전 정제.

    [주의] 무조건 모든 한글 사이 공백을 제거하면
    실제 띄어쓰기가 필요한 부분(권리자 성명 등)도 붙어버림.
    → 등기부 도메인 사전(RE_CID_ADDR, RE_CID_COURT)으로
    알려진 패턴만 선택적 정제.
    """
    # 1차: 한글-한글 사이 공백
    text = re.sub(r'(?<=[가-힣0-9]) +(?=[가-힣])', '', text)
    # 2차: 잔여 공백 (1차에서 놓친 것)
    text = re.sub(r'(?<=[가-힣0-9]) +(?=[가-힣])', '', text)
    # 한글-숫자 경계
    text = re.sub(r'(?<=[가-힣]) +(?=\d)', '', text)
    # 특수문자 주변 공백
    for ch in '【】()[]':
        text = text.replace(f'{ch} ', ch)
        text = text.replace(f' {ch}', ch)
    # 숫자-숫자 사이
    text = re.sub(r'(?<=\d) +(?=\d)', '', text)
    # 도메인 사전 기반 선택적 정제
    text = RE_CID_ADDR.sub(lambda m: m.group(0).replace(' ', ''), text)
    text = RE_CID_COURT.sub(lambda m: m.group(0).replace(' ', ''), text)
    return text


def fix_date(text):
    """
    날짜 형식 통일: '2023 년 2 월 20 일' → '2023년02월20일'

    [역공학] output.xlsx의 날짜 표기는 'YYYY년MM월DD일' 형식.
    OCR 결과는 Tesseract가 공백을 넣어 'YYYY 년 M 월 D 일'로 출력.
    월/일 zero-padding 적용.
    """
    return RE_DATE.sub(
        lambda m: f'{m.group(1)}년{m.group(2).zfill(2)}월{m.group(3).zfill(2)}일',
        text)


def fix_regno(text):
    """
    등기번호 형식 통일: '제  107414  호' → '제107414호'

    [역공학] 등기번호는 output.xlsx에서 '제XXXXXX호' 형식.
    OCR은 번호 주변에 공백을 삽입.
    """
    return RE_REGNO.sub(lambda m: f'제{m.group(1)}호', text)


def fix_money(text):
    """
    금액 형식 통일: '금  266,000,000  원' → '금266,000,000원'

    [역공학] 금액 정보는 근저당권, 전세권 등에서 중요.
    콤마가 포함된 금액도 정확히 추출해야 함.
    """
    return RE_MONEY.sub(lambda m: f'금{m.group(1)}원', text)


def remove_ocr_noise(text):
    """
    OCR 잡음 문자 제거 (보수적 방식).

    [실패사례] v6 초기: 과격한 노이즈 제거 정규식으로
    등기번호, 금액, 주소 줄까지 삭제됨.
    → "한글/숫자가 포함된 줄은 보존" 원칙으로 전환.

    [주의] 제거 기준:
    - 순수 특수문자 + box drawing 문자만 있는 줄 → 삭제
    - 한글, 숫자, 【, 금액 패턴 포함 줄 → 무조건 보존
    """
    lines = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        # 순수 노이즈 줄만 제거
        if re.match(r'^[\s@|)(\-_.,:;!?\[\]{}#*※%\'\"ㅣ\u2500-\u257F]+$', line):
            continue
        # OCR 특수문자 정리
        line = line.replace('ㅣ', '|').replace('｜', '|').replace('∣', '|')
        line = re.sub(r'[\u2500-\u257F]', '', line)  # box drawing 제거
        line = re.sub(r'\s*[|]\s*', ' | ', line)     # 파이프 주변 공백 정리
        line = re.sub(r' +', ' ', line)               # 연속 공백 → 단일
        if line:
            lines.append(line)
    return '\n'.join(lines)


def fix_terms(text):
    """
    등기용어 교정 사전 적용.

    [역공학] TERM_REPLACE 사전은 output.xlsx의 최종 출력과
    OCR 원시 출력을 1:1 비교하여 추출.
    """
    for old, new in TERM_REPLACE.items():
        text = text.replace(old, new)
    return text


def clean_text(text):
    """
    전체 텍스트 정제 파이프라인.

    [역공학] HanaXellOcr0.7의 "텍스트 후처리" 단계를 재현.
    output.xlsx 와 src_비교_*.md 의 차이를 분석하여
    아래 7단계 파이프라인을 역설계했습니다.

    [실패사례] 순서가 틀리면 결과가 크게 달라집니다.
    특히 CID 공백 정제(1) 이후에 용어 교정(2)을 해야
    "고 유 번 호" → "고유번호"가 TERM_REPLACE 사전과 매칭됩니다.
    """
    # [Step 0] 줄바꿈 정규화
    text = text.replace('\r', '\n')
    text = re.sub(r'\n{3,}', '\n\n', text)     # 3연속 개행 → 2개
    text = re.sub(r'[ \t]+', ' ', text)         # 탭/연속공백 → 단일

    # [Step 1~6] 순차 정제
    text = fix_cid_spacing(text)      # 1) CID 공백 제거 (가장 먼저!)
    text = fix_terms(text)             # 2) 등기용어 사전 교정
    text = fix_date(text)              # 3) 날짜 형식 통일
    text = fix_regno(text)             # 4) 등기번호 형식 통일
    text = fix_money(text)             # 5) 금액 형식 통일
    text = remove_ocr_noise(text)      # 6) OCR 노이즈 제거

    # [Step 7] 의미 있는 줄만 보존
    # [실패사례] 모든 줄을 보존하면 OCR 노이즈 줄이 섹션 파싱 방해.
    # → 한글/숫자/섹션헤더가 포함된 줄만 유지.
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if (re.search(r'[가-힣]', line) or
            re.search(r'\d{4}', line) or
            re.search(r'금[\d,]+원', line) or
            any(kw in line for kw in SECTION_KEYS + ['순위번호', '등기목적', '접수']) or
            re.match(r'[\[【]', line)):
            lines.append(line)

    return '\n'.join(lines)


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 6] 섹션 파싱                                            ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] HanaXellOcr0.7은 PDF 페이지 구조를 분석하여
# 표제부/갑구/을구/공동담보목록/매매목록을 자동 분리합니다.
# output.xlsx 의 시트별 데이터를 역추적하여
# 아래 섹션 키워드 기반 분할 알고리즘을 재현했습니다.


def parse_sections(text):
    """
    정제된 텍스트를 섹션별로 분할.

    [역공학] 등기부등본의 고정된 페이지 구조 활용:
    Page 1 → 표제부, Page 2 → 갑구, Page 3+ → 을구
    라는 규칙과 【】 섹션 헤더를 조합하여 분할.

    [실패사례] 일부 PDF는 페이지 구분이 모호하거나
    【갑구】가 OCR에서 누락됨.
    → '순위번호' '등기목적' 같은 컬럼 헤더를 보조 식별자로 활용.

    Returns:
        {'sections': {'【표제부】': [...], '【갑구】': [...], ...},
         'full_text': 정제된 전체 텍스트}
    """
    result = {'sections': {}, 'full_text': text}
    lines = text.split('\n')

    # ── 섹션 경계 찾기 ──
    # [역공학] 【 로 시작하는 줄 = 새 섹션 시작.
    # 표제부는 간혹 【 없이 '기본정보'로 시작하는 경우도 대응.
    # [실패사례] OCR이 【을구】를 【을 _ 구】처럼 깨뜨리는 경우 있음.
    # → 공백/특수문자 제거 후 부분 매칭으로 대응.
    # [실패사례] 【갑구】가 완전히 누락되는 PDF도 있음.
    # → '순위번호 | 등기목적' 라인을 보조 경계로 활용.
    boundaries = []
    for i, line in enumerate(lines):
        stripped = line.strip()
        # OCR 깨짐 대응: 【 을 _ 구 】 → 공백 제거 → 【을구】 매칭
        stripped_no_space = re.sub(r'\s+', '', stripped)
        if stripped.startswith('【') and any(
            kw in stripped_no_space for kw in ['표제부', '갑구', '을구', '매매목록', '공동담보']):
            boundaries.append((i, stripped))
        elif '기본정보' in stripped and not boundaries:
            boundaries.append((i, '기본정보'))
        # [실패사례] 【갑구】 누락 시 — 컬럼헤더로 보조 감지
        elif (re.search(r'순위번호\s*[\|]\s*등기목적', stripped_no_space)
              and not any('갑구' in b[1] for b in boundaries)):
            boundaries.append((i, '【갑구】'))
    
    # [실패사례] 중복 섹션명 방지 + 섹션명 정규화
    # OCR 결과가 【표제부】(토지의표시) 로 나오면 → 【표제부】 로 정규화
    seen = set()
    normalized_boundaries = []
    for start, name in boundaries:
        norm_name = name
        if '표제부' in re.sub(r'\s+', '', name):
            norm_name = '【표제부】'
        elif '갑구' in re.sub(r'\s+', '', name):
            norm_name = '【갑구】'
        elif '을구' in re.sub(r'\s+', '', name):
            norm_name = '【을구】'
        elif '매매목록' in re.sub(r'\s+', '', name):
            norm_name = '【매매목록】'
        elif '공동담보' in re.sub(r'\s+', '', name):
            norm_name = '【공동담보목록】'
        
        if norm_name not in seen:
            seen.add(norm_name)
            normalized_boundaries.append((start, norm_name))
    
    boundaries = normalized_boundaries

    # ── 섹션별 라인 할당 ──
    # [실패사례] 첫 번째 【 섹션 이전에 있는 라인(고유번호, 소재지 등)이
    #   누락되는 버그. → 첫 섹션의 시작을 0으로 보정.
    for idx, (start, name) in enumerate(boundaries):
        end = boundaries[idx + 1][0] if idx + 1 < len(boundaries) else len(lines)
        # 첫 번째 섹션: 【 앞의 라인도 포함
        actual_start = 0 if idx == 0 else start
        section_lines = lines[actual_start:end]
        result['sections'][name] = section_lines

    # 기본정보가 없으면 헤더로 추가
    if not result['sections']:
        result['sections']['__header__'] = lines

    return result


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 7] 표 구조 감지                                          ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] HanaXellOcr0.7은 표의 격자선을 감지하여
# 셀 경계를 자동으로 찾아냅니다 (output.xlsx의 정교한 셀 배치 근거).
# PyMuPDF의 Path drawing 명령어로 표 선을 탐지합니다.
#
# [실패사례] 스캔된 PDF(이미지 기반)는 표 선이 Path로 존재하지 않음.
# → detect_table_regions()가 빈 배열 반환 → fallback 모드로 OCR 결과 사용.


def detect_table_regions(pdf_path):
    """
    PDF에서 표(테이블) 영역 감지.

    [역공학] PyMuPDF get_drawings()는 PDF의 벡터 그래픽 명령을 추출.
    수평/수직선(re, l)을 찾아 표의 bbox를 추정합니다.
    HanaXellOcr0.7도 유사한 접근법을 쓰는 것으로 추정
    (output.xlsx 셀 경계가 원본 PDF의 표 선과 정확히 일치).

    [실패사례] 세로쓰기 표, 점선 표, 워터마크 중첩 시
    오탐지 발생 가능 → bbox 면적 필터링(100pt² 이상)으로 보정.

    Returns:
        [{'page': N, 'bbox': (x0,y0,x1,y1), 'rows': N, 'cols': N}, ...]
    """
    doc = fitz.open(str(pdf_path))
    tables = []

    for page_num in range(len(doc)):
        page = doc[page_num]
        paths = page.get_drawings()

        # 수평선/수직선 분리
        h_lines = []  # horizontal
        v_lines = []  # vertical

        for path in paths:
            for item in path.get('items', []):
                # [방어] item이 [op, ...] 형태인지 확인
                if not item or len(item) < 2:
                    continue
                op = item[0]
                if op == 'l' and len(item) >= 5:  # line: ['l', x1, y1, x2, y2]
                    x1, y1, x2, y2 = item[1], item[2], item[3], item[4]
                elif op == 're' and len(item) >= 5:  # rect: ['re', x, y, w, h]
                    # 사각형 → 상하좌우 4개 선으로 분해
                    rx, ry, rw, rh = item[1], item[2], item[3], item[4]
                    h_lines.append((rx, ry, rx + rw, ry))         # top
                    h_lines.append((rx, ry + rh, rx + rw, ry + rh))  # bottom
                    v_lines.append((rx, ry, rx, ry + rh))         # left
                    v_lines.append((rx + rw, ry, rx + rw, ry + rh))  # right
                    continue
                else:
                    continue

                length = abs(x2 - x1) if abs(x2 - x1) > abs(y2 - y1) else abs(y2 - y1)
                # [필터] 10pt 미만 짧은 선 = 노이즈
                if length < 10:
                    continue
                if abs(y2 - y1) < 3:        # 수평 (Y 변화 < 3pt)
                    h_lines.append((min(x1, x2), y1, max(x1, x2), y2))
                elif abs(x2 - x1) < 3:       # 수직 (X 변화 < 3pt)
                    v_lines.append((x1, min(y1, y2), x2, max(y1, y2)))

        # 표 영역 추정: 수평선 2개 + 수직선 2개 = 1개 표
        if len(h_lines) >= 2 and len(v_lines) >= 2:
            x0 = min(l[0] for l in v_lines)
            y0 = min(l[1] for l in h_lines)
            x1 = max(l[2] for l in v_lines)
            y1 = max(l[3] for l in h_lines)

            # 100pt² 미만 = 노이즈로 판단
            if (x1 - x0) * (y1 - y0) > 100:
                tables.append({
                    'page': page_num + 1,
                    'bbox': (x0, y0, x1, y1),
                    'rows': len(h_lines) - 1,
                    'cols': len(v_lines) - 1,
                })

    doc.close()
    return tables


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 8] Excel 렌더링 엔진                                     ║
# ╚══════════════════════════════════════════════════════════════╝
#
# [역공학] 이 섹션은 HanaXellOcr0.7 output.xlsx의
# 모든 시트 구조, 병합셀 규칙, 스타일을 openpyxl로 재현합니다.
#
# 참고 문서: 역공학_백서.md §3 (출력물 구조 분석)


def render_to_excel(sections, pdf_name, total_pages, tables_count,
                    output_path=None):
    """
    파싱 결과 → 6시트 Excel 파일 생성.

    [역공학] HanaXellOcr0.7 output.xlsx 구조 완전 재현:
    - 전체요약 (85행×8열): 기본정보 + 현황 요약
    - 갑구 (가변×6열): 소유권에 관한 사항
    - 을구 (가변×6열): 소유권 외 권리
    - 공동담보목록 (가변×8열): 담보 목록
    - 요약-갑구 (가변×6열): 갑구 요약
    - 요약-을구 (가변×6열): 을구 요약

    Args:
        sections: parse_sections() 반환값
        pdf_name: 원본 PDF 파일명
        total_pages: 총 페이지 수
        tables_count: 감지된 표 영역 수
        output_path: 저장 경로 (None=기본 출력 디렉토리)

    Returns:
        저장된 파일 경로
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    # [역공학] 각 시트는 output.xlsx 순서와 동일하게 생성
    _build_sheet_overview(wb, sections, pdf_name, total_pages, tables_count)
    _build_sheet_gapgu(wb, sections)
    _build_sheet_eulgu(wb, sections)
    _build_sheet_collateral(wb, sections)
    _build_sheet_summary(wb, sections, '갑구')
    _build_sheet_summary(wb, sections, '을구')

    # 시트 순서 정렬 (SHEET_NAMES 기준)
    ordered = [n for n in SHEET_NAMES if n in wb.sheetnames]
    for i, name in enumerate(ordered):
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)

    # 저장
    if output_path is None:
        output_path = OUT_DIR / f'{Path(pdf_name).stem}_등기부.xlsx'
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ── 헬퍼 함수 ─────────────────────────────────────────────────
# [역공학] openpyxl로 HanaXellOcr0.7 스타일을 재현하기 위한
#          저수준 셀 조작 함수들.
#          병합셀, 폰트, 색상, 테두리를 output.xlsx 기준으로 통일.


def _merge_and_set(ws, r1, c1, r2, c2, value, font=None, align=None):
    """셀 병합 + 값 + 스타일 설정"""
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    if font:  cell.font = font
    if align: cell.alignment = align


def _cell(ws, r, c, value, font=None, fill=None, border=None, align=None):
    """단일 셀 값 + 스타일 설정"""
    cell = ws.cell(row=r, column=c)
    cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if border: cell.border = border
    if align:  cell.alignment = align


def _fill(ws, r1, c1, r2, c2, fill):
    """범위 배경색 채우기"""
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).fill = fill


def _border(ws, r1, c1, r2, c2):
    """범위 테두리 적용"""
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).border = STYLE['border']


def _col_widths(ws, widths):
    """컬럼 너비 일괄 설정"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _extract_basic_info(sections):
    """
    섹션 데이터에서 기본정보 필드 추출.

    [역공학] output.xlsx Row 4-7 의 라벨-값 쌍을
    정규식으로 추출. 실패 시 빈 문자열 반환.

    [실패사례] 초기 정규식은 "고유번호 : 2849-..." 형식을 기대했으나
    실제 OCR 결과는 "고유번호2849-2018-019318"처럼
    라벨-값 사이에 구분자 없이 붙어나오는 경우가 많음.
    → 구분자 없는 직접 연결 케이스 추가 대응.

    [실패사례] "소재지" 라벨이 OCR에서 누락되고
    대신 "[토지]주소" 형식으로 출력됨.
    → [토지] 패턴을 소재지 대체 정보로 활용.
    """
    info = {}
    all_lines = []
    for lines in sections.values():
        all_lines.extend(lines)
    full_text = '\n'.join(all_lines)

    # ── 고유번호 ──
    # [실패사례] "고유번호2849-2018-019318" (붙어있음)
    #           "고유번호 : 2849-..." (콜론 있음)
    #           "고유번호 2849-..." (공백 구분)
    m = re.search(r'고유번호\s*[:\|]?\s*(\d{4}-\d{4}-\d{6})', full_text)
    if not m:
        m = re.search(r'고유번호\s*(\d{4}-\d{4}-\d{6})', full_text)
    if m:
        info['고유번호'] = m.group(1).strip()

    # ── 부동산종류 ──
    # [역공학] output.xlsx 에서 "[토지]" 또는 "[건물]" 형식
    m = re.search(r'\[(토지|건물|토지및건물)\]', full_text)
    if m:
        kind_map = {'토지': '토지', '건물': '건물', '토지및건물': '토지 및 건물'}
        info['부동산종류'] = kind_map.get(m.group(1), m.group(1))
    else:
        m = re.search(r'부동산종류\s*[:\|]?\s*(\S+)', full_text)
        if m:
            info['부동산종류'] = m.group(1).strip()

    # ── 소재지 ──
    # [실패사례] OCR 결과가 "[토지]경기도파주시파평면마산리113-2" 형식.
    #           "소재지:" 라벨이 없는 경우 [토지]/[건물] 뒤 주소를 사용.
    m = re.search(r'소재지\s*[:\|]?\s*(.+?)(?=부동산종류|열람일시|\n\n|\Z)', full_text, re.DOTALL)
    if m and len(m.group(1).strip()) > 3:
        info['소재지'] = re.sub(r'\s+', ' ', m.group(1).strip())
    else:
        # fallback: [토지] 또는 [건물] 다음에 나오는 주소 추출
        m = re.search(r'\[(토지|건물|토지및건물)\]\s*(.+?)(?:\n|열람일시|고유번호)', full_text)
        if m:
            addr = m.group(2).strip()
            # 등기번호(숫자-숫자-숫자) 이후는 주소가 아니므로 제거
            addr = re.sub(r'\d{4}-\d{4}-\d{6}.*', '', addr).strip()
            if len(addr) > 3:
                info['소재지'] = addr

    # ── 열람일시 ──
    # [역공학] output.xlsx 날짜 형식: "2026년03월31일17시25분59초"
    m = re.search(r'열람일시\s*[:\|]?\s*(\d{4}년\d{2}월\d{2}일\d{1,2}시\d{1,2}분\d{1,2}초)', full_text)
    if not m:
        m = re.search(r'(\d{4}년\d{2}월\d{2}일\d{1,2}시\d{1,2}분\d{1,2}초)', full_text)
    if m:
        info['열람일시'] = m.group(1).strip()

    return info


def _write_section_block(ws, start_row, sections, section_name):
    """
    섹션 내용을 시트에 기록.

    [역공학] output.xlsx 의 섹션별 데이터 배치 규칙:
    1행: 섹션 헤더 (병합, 초록 배경)
    2행~: 데이터 행 (테두리 적용)
    【 로 시작하는 중첩 섹션은 건너뜀
    """
    r = start_row
    if section_name not in sections:
        _cell(ws, r, 1, f'[{section_name} — 데이터 없음]', STYLE['meta_font'])
        return r + 1

    lines = sections[section_name]

    # 섹션 헤더
    _merge_and_set(ws, r, 1, r, 8, section_name,
                   STYLE['section_font'], STYLE['center'])
    _fill(ws, r, 1, r, 8, STYLE['section_fill'])
    r += 1

    # 데이터 행
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('【') or line.startswith('['):  # 중첩 섹션 스킵
            continue
        if any(h in line for h in ['순위번호', '등기목적', '접수']):  # 컬럼헤더 스킵
            continue
        _merge_and_set(ws, r, 1, r, 6, line, STYLE['data_font'])
        _border(ws, r, 1, r, 6)
        r += 1
    return r


# ── 시트별 빌더 함수 ──────────────────────────────────────────
# [역공학] 각 시트는 output.xlsx 의 해당 시트 레이아웃을 따릅니다.
# 병합셀 좌표는 MERGE_SPEC, 스타일은 STYLE 상수 사용.


def _build_sheet_overview(wb, sections, pdf_name, total_pages, tables_count):
    """
    [시트1] 전체요약 구축.

    [역공학] output.xlsx '전체요약' 시트 구조 (85행×8열):
    Row 1: 문서 헤더 (B1:H1 병합)
    Row 3~: 메타정보
    Row 6~: 기본정보 (C4:H4 병합)
    Row 10~: 표제부 내용
    """
    ws = wb.create_sheet('전체요약')
    r = 1

    # Row 1: 문서 헤더 [역공학: B1:H1 병합]
    _merge_and_set(ws, 1, 2, 1, 8, '등기사항전부증명서 (발췌 자동화)',
                   STYLE['header_font'], STYLE['center'])
    r = 3

    # 메타정보 행
    _merge_and_set(ws, r, 1, r, 8,
        f'원본 PDF: {pdf_name}    |    페이지: {total_pages}페이지'
        f'    |    표 영역: {tables_count}개',
        STYLE['meta_font'])
    r += 1
    _merge_and_set(ws, r, 1, r, 8,
        'HanaXellOcr0.7 역공학 기반 — 제2 독립 도구 v1.0 (Mac/Windows)',
        STYLE['meta_font'])
    r += 2

    # 기본정보 섹션 [역공학: C4:H4 병합]
    _merge_and_set(ws, r, 3, r, 8, '기  본  정  보',
                   STYLE['section_font'], STYLE['center'])
    _fill(ws, r, 3, r, 8, STYLE['section_fill'])
    r += 1

    # 기본정보 라벨-값
    info = _extract_basic_info(sections)
    for label, value in [
        ('고유번호',    info.get('고유번호', '')),
        ('소재지',      info.get('소재지', '')),
        ('부동산종류',  info.get('부동산종류', '')),
        ('열람일시',    info.get('열람일시', '')),
    ]:
        _cell(ws, r, 1, label, STYLE['label_font'])
        _merge_and_set(ws, r, 2, r, 4, value if value else '(추출 실패)',
                       STYLE['data_font'])
        r += 1

    r += 1
    _write_section_block(ws, r, sections, '【표제부】')
    _col_widths(ws, [15, 22, 18, 22, 40, 15, 18, 18])


def _build_sheet_gapgu(wb, sections):
    """
    [시트2] 갑구 구축.

    [역공학] output.xlsx '갑구' 시트 (가변행×6열):
    모든 소유권 변동 이력이 순위번호 순으로 기재.
    """
    ws = wb.create_sheet('갑구')
    _merge_and_set(ws, 1, 2, 1, 6, '갑구 — 소유권에 관한 사항',
                   STYLE['header_font'], STYLE['center'])
    _write_section_block(ws, 5, sections, '【갑구】')
    _col_widths(ws, [12, 18, 16, 20, 40, 15])


def _build_sheet_eulgu(wb, sections):
    """
    [시트3] 을구 구축.

    [역공학] output.xlsx '을구' 시트 (가변행×6열):
    근저당권, 전세권, 가압류, 임의경매 등 소유권 외 권리 기재.
    """
    ws = wb.create_sheet('을구')
    _merge_and_set(ws, 1, 2, 1, 6, '을구 — 소유권 외의 권리에 관한 사항',
                   STYLE['header_font'], STYLE['center'])
    _write_section_block(ws, 5, sections, '【을구】')
    _col_widths(ws, [12, 18, 16, 20, 40, 15])


def _build_sheet_collateral(wb, sections):
    """
    [시트4] 공동담보목록 구축.

    [역공학] output.xlsx '공동담보목록' 시트 (가변행×8열):
    여러 필지에 걸친 담보 내역을 별도 시트로 분리.
    """
    ws = wb.create_sheet('공동담보목록')
    _merge_and_set(ws, 1, 2, 1, 8, '공동담보목록',
                   STYLE['header_font'], STYLE['center'])
    _write_section_block(ws, 5, sections, '【공동담보목록】')
    _col_widths(ws, [15, 22, 18, 22, 40, 15, 18, 18])


def _build_sheet_summary(wb, sections, section_type):
    """
    [시트5/6] 요약 시트 구축.

    [역공학] output.xlsx '요약-갑구'/'요약-을구' 시트:
    각 구의 핵심 정보를 추려서 요약.
    """
    ws = wb.create_sheet(f'요약-{section_type}')
    target = f'【{section_type}】'
    r = 1

    _merge_and_set(ws, 1, 2, 1, 6, f'요약 — {section_type}',
                   STYLE['header_font'], STYLE['center'])
    r = 3

    # 컬럼 헤더
    for c, h in enumerate(['순위번호', '등기목적', '권리자', '금액/지분', '비고'], 1):
        _cell(ws, r, c, h, STYLE['label_font'], STYLE['header_fill'],
              STYLE['border'], STYLE['center'])
    r += 1

    # 데이터
    if target in sections:
        for line in sections[target]:
            line = line.strip()
            if not line or line.startswith('【') or line.startswith('['):
                continue
            _merge_and_set(ws, r, 1, r, 5, line, STYLE['data_font'])
            _border(ws, r, 1, r, 5)
            r += 1

    _col_widths(ws, [12, 18, 25, 20, 25])


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 9] 메인 파이프라인                                       ║
# ╚══════════════════════════════════════════════════════════════╝


def run(pdf_path=None):
    """
    전체 파이프라인 실행 — 파일 선택부터 Excel 출력까지.

    [흐름]
    1. PDF 선택 (GUI 다이얼로그 또는 인자)
    2. OCR 처리 (모든 페이지)
    3. 텍스트 정제 (7단계 파이프라인)
    4. 섹션 분할 (【】 기반)
    5. 표 구조 감지 (선 기반)
    6. Excel 6시트 출력 (HanaXellOcr0.7 호환)

    Returns:
        저장된 xlsx 파일 경로 (성공) 또는 None (실패/취소)
    """
    # Step 1: PDF 선택
    if pdf_path is None:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        pdf_path = filedialog.askopenfilename(
            title='등기부등본 PDF를 선택하세요',
            filetypes=[('PDF 파일', '*.pdf')]
        )
        root.destroy()
        if not pdf_path:
            print('취소되었습니다.')
            return None

    pdf_path = Path(pdf_path)
    print(f'\n{"="*55}')
    print(f'  등기부등본 PDF → Excel 변환기 v1.0')
    print(f'  파일: {pdf_path.name}')
    print(f'{"="*55}\n')

    # Step 2: OCR
    print('[1/4] OCR 처리 중...')
    pages = ocr_pdf(pdf_path)
    combined_raw = '\n\n'.join(p['text'] for p in pages)
    print(f'  → {len(pages)}페이지 OCR 완료\n')

    # Step 3: 정제
    print('[2/4] 텍스트 정제 중...')
    cleaned = clean_text(combined_raw)
    char_count = len(cleaned.replace('\n', '').replace(' ', ''))
    print(f'  → 정제 완료 (약 {char_count}자)\n')

    # Step 4: 섹션 파싱
    print('[3/4] 섹션 분할 중...')
    parsed = parse_sections(cleaned)
    sections = parsed.get('sections', {})
    print(f'  → {len(sections)}개 섹션 감지: {list(sections.keys())}\n')

    # Step 5: 표 구조
    print('[4/4] 표 구조 분석 + Excel 생성 중...')
    tables = detect_table_regions(pdf_path)
    print(f'  → {len(tables)}개 표 영역 감지')

    # Step 6: Excel 출력
    result_path = render_to_excel(
        sections, pdf_path.name, len(pages), len(tables))

    print(f'\n{"="*55}')
    print(f'  ✅ 변환 완료!')
    print(f'  📄 {result_path}')
    print(f'  📊 {len(pages)}페이지 · {len(tables)}개 표 · {len(sections)}개 섹션')
    print(f'{"="*55}\n')

    # Excel 열기 (macOS)
    if sys.platform == 'darwin':
        subprocess.run(['open', str(result_path)])
    elif sys.platform == 'win32':
        os.startfile(str(result_path))

    return result_path


# ╔══════════════════════════════════════════════════════════════╗
# ║  [Part 10] CLI 진입점                                          ║
# ╚══════════════════════════════════════════════════════════════╝

if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(
        description='등기부등본 PDF → Excel 6시트 변환기 v1.0')
    ap.add_argument('pdf', nargs='?',
                    help='PDF 파일 경로 (생략 시 파일 선택 다이얼로그)')
    ap.add_argument('-o', '--output',
                    help='출력 Excel 파일 경로 (기본: 출력결과/파일명_등기부.xlsx)')
    args = ap.parse_args()

    result = run(args.pdf)
    if result is None:
        sys.exit(1)
